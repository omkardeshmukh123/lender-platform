"""
enrich_lenders.py
=================
Single-script enrichment pipeline. No Gemini. All data from authoritative sources.

Sources (in priority order, never overwrite non-null existing values):
  1. RBI NBFC list  -- registration number, category, hq_state, hq_city (official)
  2. Screener.in    -- AUM / revenue / market cap (structured financial data)
  3. Website scraper -- phone, email, loan products, operating states

Usage:
    python enrich_lenders.py                      # enrich all lenders
    python enrich_lenders.py --approve            # enrich + auto-approve by score
    python enrich_lenders.py --approve-only       # just recompute scores + approve
    python enrich_lenders.py --only "SBFC"        # single lender (name substr)
    python enrich_lenders.py --type NBFC          # filter by company_type
    python enrich_lenders.py --missing-rbi        # only lenders without RBI reg
    python enrich_lenders.py --missing-aum        # only lenders without AUM
    python enrich_lenders.py --skip-rbi           # skip RBI enrichment
    python enrich_lenders.py --skip-screener      # skip Screener.in
    python enrich_lenders.py --skip-scraper       # skip website scraper
    python enrich_lenders.py --min-score 0.5      # approval threshold (default 0.45)
    python enrich_lenders.py --dry-run            # no DB writes
    python enrich_lenders.py --limit 50           # process only first N lenders
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import psycopg2
from psycopg2.extras import RealDictCursor
import requests

# ── env ───────────────────────────────────────────────────────
_ENV = Path(__file__).resolve().parent.parent / '.env'
try:
    from dotenv import load_dotenv
    load_dotenv(_ENV, override=False)
except ImportError:
    if _ENV.exists():
        with open(_ENV, encoding='utf-8') as _f:
            for _line in _f:
                _line = _line.strip()
                if _line and not _line.startswith('#') and '=' in _line:
                    _k, _sep, _rest = _line.partition('=')
                    _k = _k.strip(); _v = _rest.strip().strip('"').strip("'")
                    if _k and _k not in os.environ:
                        os.environ[_k] = _v

sys.path.insert(0, str(Path(__file__).parent))

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)-8s | %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
)
log = logging.getLogger(__name__)

DATABASE_URL = os.environ.get('DATABASE_URL', '')

_HEADERS = {
    'User-Agent': (
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36'
    ),
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.9',
}

ALL_INDIA_STATES = [
    "Andhra Pradesh", "Arunachal Pradesh", "Assam", "Bihar", "Chhattisgarh",
    "Goa", "Gujarat", "Haryana", "Himachal Pradesh", "Jharkhand", "Karnataka",
    "Kerala", "Madhya Pradesh", "Maharashtra", "Manipur", "Meghalaya", "Mizoram",
    "Nagaland", "Odisha", "Punjab", "Rajasthan", "Sikkim", "Tamil Nadu", "Telangana",
    "Tripura", "Uttar Pradesh", "Uttarakhand", "West Bengal", "Delhi",
    "Jammu & Kashmir", "Ladakh", "Puducherry", "Chandigarh",
    "Dadra and Nagar Haveli", "Lakshadweep", "Andaman and Nicobar Islands",
]
_STATE_SET   = set(ALL_INDIA_STATES)
_STATE_LOWER = {s.lower(): s for s in ALL_INDIA_STATES}

VALID_PRODUCTS = [
    "MSME Loan", "Personal Loan", "Home Loan", "Business Loan",
    "Vehicle Loan", "Gold Loan", "Education Loan", "Micro Loan",
    "Loan Against Property", "Working Capital", "Agriculture Loan",
    "Credit Card", "Consumer Durable Loan", "EV Loan", "Two Wheeler Loan",
    "Microfinance", "Rural Loan", "Supply Chain Finance",
]

VALID_RBI_CATEGORIES = [
    "NBFC-D", "NBFC-ND", "NBFC-ND-SI", "NBFC-MFI", "NBFC-IFC",
    "NBFC-Factor", "NBFC-ICC", "NBFC-P2P", "NBFC-AA", "NBFC-IC",
]

# Fields scored for quality_score
_QUALITY_FIELDS = [
    'company_name', 'website', 'hq_state', 'company_type',
    'primary_loan_segments', 'operating_states', 'phone', 'email',
    'aum_crores', 'established_year', 'employee_count',
    'rbi_registration_number', 'rbi_category', 'operating_intensity', 'hq_location',
    'cin',  # MCA21: company identification number
]


# ─────────────────────────────────────────────────────────────
# DB
# ─────────────────────────────────────────────────────────────

def get_db():
    if not DATABASE_URL:
        raise RuntimeError('DATABASE_URL not set in .env')
    return psycopg2.connect(DATABASE_URL, connect_timeout=15)


def fetch_lenders(conn, filters: Dict) -> List[Dict]:
    """Fetch lenders from DB with optional filters."""
    conditions = ['1=1']
    params: List[Any] = []

    if filters.get('only'):
        conditions.append('LOWER(company_name) LIKE %s')
        params.append(f'%{filters["only"].lower()}%')
    if filters.get('type'):
        conditions.append('company_type = %s')
        params.append(filters['type'])
    if filters.get('missing_rbi'):
        conditions.append('rbi_registration_number IS NULL')
    if filters.get('missing_aum'):
        conditions.append('aum_crores IS NULL')

    where = ' AND '.join(conditions)
    limit_clause = f'LIMIT {int(filters["limit"])}' if filters.get('limit') else ''

    sql = f"""
        SELECT id, company_name, company_type, website,
               aum_crores, aum_category, last_year_revenue,
               is_listed, stock_symbol,
               primary_loan_segments, primary_product,
               hq_location, hq_state, operating_states, operating_intensity, pan_india,
               rbi_category, rbi_registration_number,
               established_year, employee_count, branch_count,
               phone, email, rural_presence,
               cin, company_status, authorized_capital_lakhs, paid_up_capital_lakhs,
               mca21_status,
               quality_score, approval_status
        FROM lenders
        WHERE {where}
        ORDER BY company_name
        {limit_clause}
    """
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(sql, params)
        return [dict(r) for r in cur.fetchall()]


def _is_empty(val: Any) -> bool:
    if val is None:
        return True
    if isinstance(val, list) and len(val) == 0:
        return True
    if isinstance(val, str) and val.strip() == '':
        return True
    return False


def _compute_quality(record: Dict) -> float:
    present = sum(1 for f in _QUALITY_FIELDS if not _is_empty(record.get(f)))
    return round(present / len(_QUALITY_FIELDS), 3)


def upsert_enriched(conn, record: Dict, dry_run: bool = False) -> str:
    """
    UPDATE only the enrichable fields for an existing lender.
    Never touches approval_status.
    Returns 'updated' | 'dry_run'.
    """
    if dry_run:
        return 'dry_run'

    fields = [
        'aum_crores', 'aum_category', 'last_year_revenue',
        'is_listed', 'stock_symbol',
        'primary_loan_segments', 'primary_product',
        'hq_location', 'hq_state', 'operating_states', 'operating_intensity', 'pan_india',
        'rbi_category', 'rbi_registration_number',
        'established_year', 'employee_count', 'branch_count',
        'phone', 'email', 'rural_presence',
        'cin', 'company_status', 'authorized_capital_lakhs', 'paid_up_capital_lakhs',
        'quality_score', 'data_source', 'extraction_status',
    ]
    # Only update fields that are non-null in the enriched record
    update_pairs = []
    values = []
    for f in fields:
        val = record.get(f)
        if not _is_empty(val):
            update_pairs.append(f'{f} = %s')
            values.append(_coerce(f, val))

    # mca21_status and mca21_enriched_at are always written when present
    if record.get('mca21_status') and record['mca21_status'] != 'not_enriched':
        update_pairs.append('mca21_status = %s')
        values.append(record['mca21_status'])
        update_pairs.append('mca21_enriched_at = NOW()')

    if not update_pairs:
        return 'no_change'

    update_pairs.append('last_scraped_at = NOW()')
    sql = f"UPDATE lenders SET {', '.join(update_pairs)} WHERE id = %s"
    values.append(record['id'])

    with conn.cursor() as cur:
        cur.execute(sql, values)
    conn.commit()
    return 'updated'


def insert_lender(conn, record: Dict, dry_run: bool = False) -> str:
    """Insert a lender that doesn't exist in DB yet."""
    if dry_run:
        return 'dry_run'

    cols = [
        'company_name', 'company_type', 'website',
        'aum_crores', 'aum_category', 'last_year_revenue',
        'is_listed', 'primary_loan_segments', 'primary_product',
        'hq_location', 'hq_state', 'operating_states', 'operating_intensity', 'pan_india',
        'rbi_category', 'rbi_registration_number',
        'established_year', 'employee_count', 'branch_count',
        'phone', 'email', 'rural_presence',
        'quality_score', 'data_source', 'extraction_status', 'approval_status',
    ]
    values = [_coerce(c, record.get(c)) for c in cols]
    placeholders = ', '.join(['%s'] * len(cols)) + ', NOW()'
    col_list = ', '.join(cols) + ', last_scraped_at'

    sql = f"""
        INSERT INTO lenders ({col_list})
        VALUES ({placeholders})
        ON CONFLICT (company_name) DO NOTHING
    """
    with conn.cursor() as cur:
        cur.execute(sql, values)
    conn.commit()
    return 'inserted'


_TEXT_ARRAY = {'primary_loan_segments', 'operating_states'}
_BOOL_FIELDS = {'is_listed', 'pan_india', 'rural_presence'}
_FLOAT_FIELDS = {'aum_crores', 'last_year_revenue', 'quality_score',
                 'authorized_capital_lakhs', 'paid_up_capital_lakhs'}
_INT_FIELDS   = {'established_year', 'employee_count', 'branch_count'}


def _coerce(field: str, val: Any) -> Any:
    if val is None or val == '' or val == 'None':
        return None
    if field in _TEXT_ARRAY:
        if isinstance(val, list):
            return val
        try:
            return json.loads(val)
        except Exception:
            return []
    if field in _BOOL_FIELDS:
        if isinstance(val, bool):
            return val
        return str(val).lower() in ('true', '1', 'yes')
    if field in _FLOAT_FIELDS:
        try:
            return float(val)
        except (TypeError, ValueError):
            return None
    if field in _INT_FIELDS:
        try:
            return int(float(val))
        except (TypeError, ValueError):
            return None
    return val


def bulk_update_quality_and_approve(
    conn, min_score: float, dry_run: bool = False
) -> Tuple[int, int]:
    """
    Recompute quality_score for all lenders and approve those >= min_score.
    Returns (scores_updated, newly_approved).
    """
    if dry_run:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT COUNT(*) FROM lenders
                WHERE approval_status != 'approved'
                AND (
                    (hq_state IS NOT NULL) AND
                    (primary_loan_segments IS NOT NULL AND array_length(primary_loan_segments,1) > 0)
                )
            """)
            would_approve = cur.fetchone()[0]
        log.info('[dry-run] would approve ~%d lenders', would_approve)
        return 0, would_approve

    with conn.cursor() as cur:
        # Approve lenders that have minimum data
        cur.execute("""
            UPDATE lenders
            SET approval_status = 'approved'
            WHERE approval_status != 'approved'
            AND hq_state IS NOT NULL
            AND primary_loan_segments IS NOT NULL
            AND array_length(primary_loan_segments, 1) > 0
            AND quality_score >= %s
        """, (min_score,))
        newly_approved = cur.rowcount
    conn.commit()
    return 0, newly_approved


# ─────────────────────────────────────────────────────────────
# SOURCE 1: RBI NBFC LIST
# ─────────────────────────────────────────────────────────────

_rbi_entries: Optional[List[Dict]] = None  # module-level cache

_RBI_LOCAL_PATHS = [
    Path(__file__).parent.parent / 'data' / 'input' / 'rbi_nbfc_list.xlsx',
    Path(__file__).parent.parent / 'data' / 'input' / 'rbi_nbfc_list.xls',
    Path(__file__).parent.parent / 'data' / 'input' / 'rbi_nbfc_list.csv',
]


def _parse_rbi_table(headers: List[str], data_rows: List[List[str]]) -> List[Dict]:
    """Convert raw header + row data into normalised entry dicts."""
    col: Dict[str, int] = {}
    for i, h in enumerate(headers):
        h_low = str(h).lower().strip()
        if ('name' in h_low or 'company' in h_low) and 'name' not in col:
            col['name'] = i
        elif 'registr' in h_low and 'reg_no' not in col:
            col['reg_no'] = i
        elif 'categor' in h_low and 'category' not in col:
            col['category'] = i
        elif 'state' in h_low and 'state' not in col:
            col['state'] = i
        elif ('district' in h_low or 'city' in h_low) and 'city' not in col:
            col['city'] = i

    entries = []
    for row in data_rows:
        if not row:
            continue
        entry: Dict[str, str] = {}
        for field, idx in col.items():
            if idx < len(row):
                entry[field] = str(row[idx]).strip()
        if entry.get('name') and entry['name'].lower() not in ('nan', 'none', ''):
            entries.append(entry)
    return entries


def _load_rbi_from_file() -> List[Dict]:
    """Try to load the RBI NBFC list from a locally downloaded file."""
    for path in _RBI_LOCAL_PATHS:
        if not path.exists():
            continue
        log.info('[rbi] loading from local file: %s', path.name)
        try:
            if path.suffix == '.csv':
                import csv
                with open(path, encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                if len(rows) < 2:
                    continue
                return _parse_rbi_table(rows[0], rows[1:])
            else:
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                    ws = wb.active
                    all_rows = [[str(c.value or '') for c in row] for row in ws.iter_rows()]
                except Exception:
                    import xlrd
                    wb = xlrd.open_workbook(path)
                    ws = wb.sheet_by_index(0)
                    all_rows = [[str(ws.cell(r, c).value) for c in range(ws.ncols)]
                                for r in range(ws.nrows)]
                if len(all_rows) < 2:
                    continue
                return _parse_rbi_table(all_rows[0], all_rows[1:])
        except Exception as exc:
            log.warning('[rbi] could not read %s: %s', path.name, exc)

    return []


def _load_rbi_from_web() -> List[Dict]:
    """
    Scrape the RBI NBFC list via HTTP.
    The RBI site uses bot detection; returns empty list on failure with
    instructions for manual download.
    """
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        log.warning('[rbi] pip install beautifulsoup4 to enable RBI enrichment')
        return []

    url = 'https://www.rbi.org.in/Scripts/BS_NBFCList.aspx'
    session = requests.Session()

    # Use realistic browser headers to reduce bot detection
    headers = {
        **_HEADERS,
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection':      'keep-alive',
        'Cache-Control':   'no-cache',
        'Upgrade-Insecure-Requests': '1',
        'Sec-Fetch-Site':  'none',
        'Sec-Fetch-Mode':  'navigate',
        'Sec-Fetch-User':  '?1',
        'Sec-Fetch-Dest':  'document',
    }

    try:
        resp = session.get(url, headers=headers, timeout=25)
    except Exception as exc:
        log.warning('[rbi] network error: %s', exc)
        return []

    if resp.status_code == 418:
        log.warning(
            '[rbi] RBI site returned 418 (bot detection).\n'
            '  To enable RBI enrichment:\n'
            '  1. Open https://www.rbi.org.in/Scripts/BS_NBFCList.aspx in a browser\n'
            '  2. Click "Download" to save the Excel file\n'
            '  3. Save it as:  data/input/rbi_nbfc_list.xlsx\n'
            '  4. Re-run this script'
        )
        return []

    if not resp.ok:
        log.warning('[rbi] GET failed: HTTP %s', resp.status_code)
        return []

    soup = BeautifulSoup(resp.text, 'html.parser')

    # Extract __VIEWSTATE and try POST
    vs_tag = soup.find('input', {'name': '__VIEWSTATE'})
    ev_tag = soup.find('input', {'name': '__EVENTVALIDATION'})

    if vs_tag and ev_tag:
        post_data = {
            '__VIEWSTATE':       vs_tag.get('value', ''),
            '__EVENTVALIDATION': ev_tag.get('value', ''),
            '__EVENTTARGET':     '',
            '__EVENTARGUMENT':   '',
            'ddlRegulator':      '0',
            'ddlCategory':       '0',
            'ddlState':          '0',
            'btnSubmit':         'Search',
        }
        try:
            resp2 = session.post(url, data=post_data, headers=headers, timeout=30)
            if resp2.ok:
                soup = BeautifulSoup(resp2.text, 'html.parser')
        except Exception:
            pass  # fall through to parse whatever we have

    # Parse any table with Company + Registration columns
    for t in soup.find_all('table'):
        text_preview = t.get_text()[:400].lower()
        if 'company' in text_preview and ('registr' in text_preview or 'category' in text_preview):
            rows = t.find_all('tr')
            if len(rows) < 2:
                continue
            headers_row = [cell.get_text(strip=True) for cell in rows[0].find_all(['th', 'td'])]
            data_rows   = [
                [cell.get_text(strip=True) for cell in row.find_all('td')]
                for row in rows[1:]
            ]
            entries = _parse_rbi_table(headers_row, data_rows)
            if entries:
                return entries

    log.warning('[rbi] could not parse NBFC table from page')
    return []


def load_rbi_nbfc_list() -> List[Dict]:
    """Load the RBI NBFC list. Tries local file first, then web. Cached."""
    global _rbi_entries
    if _rbi_entries is not None:
        return _rbi_entries

    # Local file takes priority (avoids bot detection, always up-to-date on demand)
    entries = _load_rbi_from_file()
    if not entries:
        entries = _load_rbi_from_web()

    log.info('[rbi] loaded %d entries total', len(entries))
    _rbi_entries = entries
    return _rbi_entries


def _token_similarity(a: str, b: str) -> float:
    """Token overlap score between two strings."""
    def tokens(s: str) -> set:
        return set(re.sub(r'[^a-z0-9]', ' ', s.lower()).split())
    ta, tb = tokens(a), tokens(b)
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / max(len(ta), len(tb))


def rbi_lookup(name: str, threshold: float = 0.55) -> Dict[str, Optional[str]]:
    """Find best RBI match for a company name. Returns enrichment dict."""
    entries = load_rbi_nbfc_list()
    if not entries:
        return {}

    best_score = 0.0
    best = None
    for entry in entries:
        score = _token_similarity(name, entry.get('name', ''))
        if score > best_score:
            best_score = score
            best = entry

    if best_score < threshold or not best:
        return {}

    log.debug('  [rbi] "%s" -> "%s" (%.2f)', name, best.get('name'), best_score)

    result: Dict[str, Optional[str]] = {}

    # Registration number — reject CIN format (MCA number, not RBI CoR)
    reg = (best.get('reg_no') or '').strip()
    _cin = re.compile(r'^[LUu]\d{5}[A-Z]{2}\d{4}[A-Z]{3}\d{6}$')
    if reg and any(c.isdigit() for c in reg) and not _cin.match(reg):
        result['rbi_registration_number'] = reg

    # Category
    cat_raw = (best.get('category') or '').strip().upper().replace(' ', '-')
    for allowed in VALID_RBI_CATEGORIES:
        if allowed.upper() in cat_raw or cat_raw in allowed.upper():
            result['rbi_category'] = allowed
            break

    # Location
    city = (best.get('city') or '').strip()
    if city:
        result['hq_location'] = city

    state_raw = (best.get('state') or '').strip()
    if state_raw:
        normalized = _STATE_LOWER.get(state_raw.lower()) or (
            state_raw if state_raw in _STATE_SET else None
        )
        if normalized:
            result['hq_state'] = normalized

    return result


# ─────────────────────────────────────────────────────────────
# SOURCE 2: SCREENER.IN
# ─────────────────────────────────────────────────────────────

_screener_csrf: Optional[str] = None  # module-level CSRF token cache


def _get_screener_csrf(session: requests.Session) -> str:
    """Fetch Screener homepage to get a session cookie + CSRF token."""
    global _screener_csrf
    if _screener_csrf:
        return _screener_csrf
    try:
        resp = session.get(
            'https://www.screener.in/',
            headers=_HEADERS,
            timeout=15,
        )
        # CSRF token is in a meta tag or cookie
        token = session.cookies.get('csrftoken', '')
        if not token:
            import re as _re
            m = _re.search(r'csrfmiddlewaretoken.*?value=["\']([^"\']+)', resp.text)
            if m:
                token = m.group(1)
        _screener_csrf = token
        return token
    except Exception:
        return ''


def _screener_search(name: str, session: requests.Session) -> Optional[str]:
    """
    Find the Screener.in company slug for a given name.
    Returns the slug string (e.g. 'SBFC') or None.
    """
    try:
        csrf = _get_screener_csrf(session)
        search_headers = {
            **_HEADERS,
            'Accept': 'application/json',
            'Referer': 'https://www.screener.in/',
            'X-Requested-With': 'XMLHttpRequest',
        }
        if csrf:
            search_headers['X-CSRFToken'] = csrf

        resp = session.get(
            'https://www.screener.in/api/company/search/',
            params={'q': name, 'v': '3'},
            headers=search_headers,
            timeout=15,
        )
        if resp.status_code == 403:
            log.debug('  [screener] 403 on search — CSRF issue')
            return None
        if not resp.ok:
            return None
        data = resp.json()
        if not data:
            return None
        # Pick best name match
        best_score = 0.0
        best_slug  = None
        for item in data[:5]:
            score = _token_similarity(name, item.get('name', ''))
            if score > best_score:
                best_score = score
                url_path = item.get('url', '')
                # url is like /company/SBFC/ or /company/SBFC/consolidated/
                parts = [p for p in url_path.strip('/').split('/') if p]
                if len(parts) >= 2:
                    best_slug = parts[1]
        if best_score >= 0.5:
            return best_slug
        return None
    except Exception as exc:
        log.debug('  [screener] search error: %s', exc)
        return None


def _parse_screener_number(text: str) -> Optional[float]:
    """Parse numbers like '₹4,943 Cr', '14,200', '1.2 Lakh Cr' etc."""
    if not text:
        return None
    text = text.replace(',', '').replace('₹', '').replace('Rs.', '').strip()

    # Lakh Cr → multiply
    m = re.search(r'([\d.]+)\s*lakh\s*cr', text, re.I)
    if m:
        return float(m.group(1)) * 1_00_000

    # Thousand Cr
    m = re.search(r'([\d.]+)\s*thousand\s*cr', text, re.I)
    if m:
        return float(m.group(1)) * 1000

    # Cr directly
    m = re.search(r'([\d.]+)\s*cr', text, re.I)
    if m:
        return float(m.group(1))

    # Plain number
    m = re.search(r'([\d.]+)', text)
    if m:
        return float(m.group(1))

    return None


def screener_enrich(name: str, session: requests.Session) -> Dict[str, Any]:
    """
    Scrape Screener.in for financial data.
    Returns dict with aum_crores, last_year_revenue, is_listed, established_year etc.
    """
    result: Dict[str, Any] = {}

    slug = _screener_search(name, session)
    if not slug:
        log.debug('  [screener] no match for "%s"', name)
        return result

    try:
        from bs4 import BeautifulSoup
    except ImportError:
        return result

    try:
        resp = session.get(
            f'https://www.screener.in/company/{slug}/',
            headers={**_HEADERS, 'Referer': 'https://www.screener.in/'},
            timeout=20,
        )
        if not resp.ok:
            log.debug('  [screener] company page HTTP %s for "%s"', resp.status_code, slug)
            return result
    except Exception as exc:
        log.debug('  [screener] page error: %s', exc)
        return result

    soup = BeautifulSoup(resp.text, 'html.parser')

    # ── is_listed: Screener.in only has listed companies ──────
    result['is_listed'] = True
    result['stock_symbol'] = slug  # Screener slug = NSE/BSE ticker

    # ── Parse financial tables ─────────────────────────────────
    # Screener tables: col 0 = label, cols 1..N = years oldest→newest
    # We always want the LAST column (most recent).

    def _last_val(row_cells: Any) -> Optional[float]:
        """Get numeric value from the last (most recent) data column."""
        cells = list(row_cells)
        # col 0 is label; find last non-empty data cell from the right
        for cell in reversed(cells[1:]):
            text = cell.get_text(strip=True).replace(',', '')
            if text and text not in ('-', ''):
                return _parse_screener_number(text)
        return None

    revenue      = None
    total_assets = None

    for table in soup.find_all('table'):
        table_text = table.get_text().lower()
        rows = table.find_all('tr')
        if len(rows) < 2:
            continue

        for row in rows[1:]:  # skip header row
            cells = row.find_all(['th', 'td'])
            if not cells:
                continue
            label = cells[0].get_text(strip=True).lower().rstrip('+').strip()

            # Revenue from annual P&L
            # Screener annual table header pattern: "Mar 2014, Mar 2015..." (same month every year)
            # Quarterly pattern: "Dec 2022, Mar 2023, Jun 2023..." (different months)
            if revenue is None and label in (
                'revenue', 'net revenue', 'revenue from operations',
                'sales', 'net sales', 'total revenue',
            ):
                # Detect quarterly: header months vary (not all the same)
                header_cells = rows[0].find_all(['th', 'td'])
                month_names = re.findall(
                    r'(jan|feb|apr|may|jun|jul|aug|oct|nov|dec)',
                    ' '.join(h.get_text() for h in header_cells), re.I
                )
                if month_names:
                    continue  # quarterly table — skip (annual only has Mar)
                val = _last_val(cells)
                if val and val > 0:
                    revenue = val

            # Total Assets from balance sheet
            if total_assets is None and label == 'total assets':
                val = _last_val(cells)
                if val and val > 0:
                    total_assets = val

    if revenue:
        result['last_year_revenue'] = revenue
    if total_assets:
        result['aum_crores'] = total_assets  # total assets ≈ loan book for NBFCs/banks

    # ── Established year from About section ────────────────────
    about_section = soup.find(id=re.compile(r'about', re.I)) or \
                    soup.find('div', class_=re.compile(r'about', re.I))
    if about_section:
        about_text = about_section.get_text()
        m = re.search(r'(?:incorporated|established|founded|since)\s+(?:in\s+)?(\d{4})',
                      about_text, re.I)
        if m:
            yr = int(m.group(1))
            if 1850 <= yr <= 2025:
                result['established_year'] = yr

    filled = {k: v for k, v in result.items() if v not in (None, '', [], False)}
    if filled:
        log.info('  [screener] "%s" (slug=%s) filled: %s', name, slug,
                 list(filled.keys()))
    else:
        log.debug('  [screener] no useful data for "%s"', name)

    return result


# ─────────────────────────────────────────────────────────────
# SOURCE 3: MCA21 (offline CSV-first, web fallback)
# ─────────────────────────────────────────────────────────────
# Primary:  data/input/mca21_companies.csv  (bulk MCA21 data)
#           Download from: https://www.mca.gov.in/mcafoportal/viewCompanyMasterData.do
#           Columns expected: CIN, company_name, company_status,
#                             date_of_incorporation, state, authorised_capital,
#                             paidup_capital, email, listing_status
#
# Fallback: CIN extraction from website URL (vakilsearch/charteredone URLs
#           embed the CIN, which gives us state + year + listing status
#           with zero API calls). Use sync_nbfc_csv.py for bulk CIN extraction.
#
# Merge rule: NEVER overwrites non-null existing DB values.

_MCA21_CSV  = Path(__file__).parent.parent / 'data' / 'input' / 'mca21_companies.csv'
_CIN_RE     = re.compile(r'\b([LU]\d{5}[A-Z]{2}\d{4}[A-Z]{2,3}\d{6})\b')
_CIN_STRICT = re.compile(r'^[LU]\d{5}[A-Z]{2}\d{4}[A-Z]{3}\d{6}$')

# MCA state code → canonical state name (same map as sync_nbfc_csv.py)
_MCA_STATE_CODES: Dict[str, str] = {
    'AN': 'Andaman and Nicobar Islands', 'AP': 'Andhra Pradesh',
    'AR': 'Arunachal Pradesh', 'AS': 'Assam', 'BR': 'Bihar',
    'CH': 'Chandigarh', 'CG': 'Chhattisgarh', 'CT': 'Chhattisgarh',
    'DD': 'Dadra and Nagar Haveli', 'DL': 'Delhi',
    'DN': 'Dadra and Nagar Haveli', 'GA': 'Goa', 'GJ': 'Gujarat',
    'HR': 'Haryana', 'HP': 'Himachal Pradesh', 'JK': 'Jammu & Kashmir',
    'JH': 'Jharkhand', 'KA': 'Karnataka', 'KL': 'Kerala',
    'LD': 'Lakshadweep', 'MP': 'Madhya Pradesh', 'MH': 'Maharashtra',
    'MN': 'Manipur', 'ML': 'Meghalaya', 'MZ': 'Mizoram',
    'NL': 'Nagaland', 'OD': 'Odisha', 'OR': 'Odisha',
    'PY': 'Puducherry', 'PB': 'Punjab', 'RJ': 'Rajasthan',
    'SK': 'Sikkim', 'TN': 'Tamil Nadu', 'TG': 'Telangana',
    'TR': 'Tripura', 'UP': 'Uttar Pradesh', 'UK': 'Uttarakhand',
    'UT': 'Uttarakhand', 'WB': 'West Bengal',
}

# Module-level CSV cache
_mca21_csv_cache: Optional[List[Dict]] = None
_mca21_by_name:   Optional[Dict[str, Dict]] = None


def _extract_year_from_date(date_str: str) -> Optional[int]:
    """Extract 4-digit year from strings like '23 Apr 1987', '1987-04-23'."""
    if not date_str:
        return None
    m = re.search(r'\b(1[89]\d{2}|20[012]\d)\b', date_str)
    if m:
        yr = int(m.group(1))
        return yr if 1850 <= yr <= 2026 else None
    return None


def _parse_capital_lakhs(text: str) -> Optional[float]:
    """
    Parse capital amounts from MCA21 CSV strings.
    MCA bulk CSV stores amounts in actual rupees (not lakhs); we convert.
    Examples: '1,00,00,000' → 1000.0 Lakhs, '50,00,000' → 50.0 Lakhs
    """
    if not text:
        return None
    cleaned = re.sub(r'[₹Rs.,\s]', '', str(text))
    try:
        rupees = float(cleaned)
        return round(rupees / 1e5, 2)
    except (ValueError, TypeError):
        return None


def _parse_cin_fields(cin: str) -> Dict[str, Any]:
    """
    Decode a CIN into state and year.
    CIN: L/U + 5-digit NIC + 2-letter state + 4-digit year + class + 6-digit seq
    """
    result: Dict[str, Any] = {}
    if not cin or len(cin) < 12:
        return result

    # Listing status
    result['is_listed'] = cin[0].upper() == 'L'

    # State code (characters at index 6-7 in standard CIN)
    # e.g., L65910MH1994PLC077895 → MH at position 6
    state_code = cin[6:8].upper()
    state = _MCA_STATE_CODES.get(state_code)
    if state:
        result['hq_state'] = state

    # Year (characters 8-11)
    year_str = cin[8:12]
    try:
        yr = int(year_str)
        if 1850 <= yr <= 2026:
            result['established_year'] = yr
    except ValueError:
        pass

    return result


def _load_mca21_csv() -> Optional[Dict[str, Dict]]:
    """
    Load the MCA21 bulk CSV into memory (cached).
    Returns dict keyed by normalised company name, or None if file absent.

    Expected CSV columns (case-insensitive):
        CIN, company_name, company_status, date_of_incorporation,
        state, authorised_capital, paidup_capital, email, listing_status
    """
    global _mca21_csv_cache, _mca21_by_name

    if _mca21_by_name is not None:
        return _mca21_by_name

    if not _MCA21_CSV.exists():
        log.info(
            '[mca21] Bulk CSV not found at %s\n'
            '  To enable MCA21 enrichment from CSV:\n'
            '  1. Visit https://www.mca.gov.in/mcafoportal/viewCompanyMasterData.do\n'
            '  2. Download the XLSX/CSV and save as data/input/mca21_companies.csv\n'
            '  3. Re-run this script\n'
            '  (Alternatively, run: python backend/sync_nbfc_csv.py)',
            _MCA21_CSV,
        )
        _mca21_by_name = {}  # empty, not None — avoid repeated log
        return _mca21_by_name

    import csv as _csv

    _mca21_csv_cache = []
    _mca21_by_name   = {}

    def _norm(s: str) -> str:
        s = s.lower()
        s = re.sub(r'[^\w\s]', ' ', s)
        s = re.sub(r'\s+', ' ', s).strip()
        return s

    try:
        with open(_MCA21_CSV, encoding='utf-8-sig', newline='') as f:
            reader = _csv.DictReader(f)
            headers_lower = None
            for raw_row in reader:
                # Normalise header names on first row
                if headers_lower is None:
                    headers_lower = {k.strip().lower(): k for k in raw_row}
                row = {k.strip().lower(): v.strip() for k, v in raw_row.items()}
                name = row.get('company_name', '').strip()
                if not name or name.lower() in ('nan', 'none', ''):
                    continue
                _mca21_csv_cache.append(row)
                _mca21_by_name[_norm(name)] = row

        log.info('[mca21] loaded %d companies from bulk CSV', len(_mca21_csv_cache))
    except Exception as exc:
        log.warning('[mca21] CSV load error: %s', exc)
        _mca21_by_name = {}

    return _mca21_by_name


def _mca21_lookup(name: str, threshold: float = 0.55) -> Optional[Dict]:
    """Look up a company in the MCA21 CSV by fuzzy name match."""
    index = _load_mca21_csv()
    if not index:
        return None

    def _norm(s: str) -> str:
        s = s.lower(); s = re.sub(r'[^\w\s]', ' ', s)
        return re.sub(r'\s+', ' ', s).strip()

    _STOP = {'limited', 'ltd', 'private', 'pvt', 'financial', 'finance',
             'capital', 'services', 'solutions', 'company', 'india', 'the',
             'and', 'of', 'a', 'an'}

    def _tokens(s: str) -> set:
        return {t for t in _norm(s).split() if t not in _STOP and len(t) >= 3}

    query_tokens = _tokens(name)
    if not query_tokens:
        return None

    # Exact normalised name first
    norm_name = _norm(name)
    if norm_name in index:
        return index[norm_name]

    # Token overlap search
    best_ratio = 0.0
    best_row   = None
    for reg_norm, row in index.items():
        reg_tokens = _tokens(reg_norm)
        if not reg_tokens:
            continue
        overlap = len(query_tokens & reg_tokens)
        ratio   = overlap / max(len(query_tokens), len(reg_tokens))
        if ratio > best_ratio:
            best_ratio = ratio
            best_row   = row

    return best_row if best_ratio >= threshold else None


def mca21_enrich(name: str, website: str = '',
                 session: Optional[Any] = None) -> Dict[str, Any]:
    """
    Enrich a lender with MCA21 company registration data.

    Strategy (in order):
      1. Look up in offline MCA21 bulk CSV (if available)
      2. Extract CIN from the lender's website URL (vakilsearch/charteredone)
      3. Decode CIN → state, year, listing status (zero API calls)

    Returns dict with any subset of:
      cin, established_year, company_status, hq_state,
      authorized_capital_lakhs, paid_up_capital_lakhs, is_listed,
      mca21_status ('enriched' | 'not_found')
    """
    result: Dict[str, Any] = {}

    # ── Strategy 1: Offline CSV lookup ────────────────────────
    csv_row = _mca21_lookup(name)
    if csv_row:
        cin = csv_row.get('cin', '').strip().upper()
        if cin and _CIN_STRICT.match(cin):
            result['cin'] = cin
            result.update(_parse_cin_fields(cin))

        # Company status
        raw_status = csv_row.get('company_status', '').lower().strip()
        if 'active' in raw_status:
            result['company_status'] = 'active'
        elif 'struck' in raw_status:
            result['company_status'] = 'struck_off'
        elif 'dormant' in raw_status:
            result['company_status'] = 'dormant'
        elif 'dissolved' in raw_status or 'wound' in raw_status:
            result['company_status'] = 'dissolved'
        elif raw_status:
            result['company_status'] = raw_status[:30]

        # Incorporation date → year (overrides CIN-derived year if present)
        yr = _extract_year_from_date(csv_row.get('date_of_incorporation', ''))
        if yr:
            result['established_year'] = yr

        # State from registered office (more precise than CIN code)
        state_raw = csv_row.get('state', '').strip()
        if state_raw:
            norm = _STATE_LOWER.get(state_raw.lower())
            if norm:
                result['hq_state'] = norm

        # Capital
        auth_cap = _parse_capital_lakhs(csv_row.get('authorised_capital', ''))
        if auth_cap:
            result['authorized_capital_lakhs'] = auth_cap

        paid_cap = _parse_capital_lakhs(csv_row.get('paidup_capital', ''))
        if paid_cap:
            result['paid_up_capital_lakhs'] = paid_cap

        # Listing
        listing_raw = csv_row.get('listing_status', '').lower()
        if 'listed' in listing_raw and 'unlisted' not in listing_raw:
            result['is_listed'] = True
        elif 'unlisted' in listing_raw:
            result['is_listed'] = False

        result['mca21_status'] = 'enriched'

    # ── Strategy 2: CIN from website URL ──────────────────────
    if not result.get('cin') and website:
        m = _CIN_RE.search(website)
        if m:
            cin = m.group(1)
            result['cin'] = cin
            result.update(_parse_cin_fields(cin))
            result['mca21_status'] = 'enriched'

    if not result.get('mca21_status'):
        result['mca21_status'] = 'not_found'
        return result

    filled = {k: v for k, v in result.items()
              if v not in (None, '', []) and k != 'mca21_status'}
    if filled:
        log.info('  [mca21] "%s" filled: %s', name, list(filled.keys()))

    return result


# ─────────────────────────────────────────────────────────────
# SOURCE 4: WEBSITE SCRAPER
# ─────────────────────────────────────────────────────────────

def scraper_enrich(name: str, website: str) -> Optional[Dict]:
    """Run SingleLenderScraper on the lender website."""
    if not website:
        return None
    try:
        from scraper.lender_scraper import SingleLenderScraper
        result = SingleLenderScraper().scrape(name, website)
        filled = sum(
            1 for k, v in result.items()
            if not k.startswith('_') and v not in (None, '', [], False)
        )
        log.info('  [scraper] %d fields extracted', filled)
        return result
    except Exception as exc:
        log.warning('  [scraper] failed for "%s": %s', name, exc)
        return None


# ─────────────────────────────────────────────────────────────
# MERGE
# ─────────────────────────────────────────────────────────────

def merge(
    db:       Dict,
    rbi:      Optional[Dict],
    mca21:    Optional[Dict],
    screener: Optional[Dict],
    scraper:  Optional[Dict],
) -> Dict:
    """
    Merge all source data onto the DB record.
    NEVER overwrites a non-null existing DB value.

    Priority for each field (highest → lowest):
      existing_db > rbi_official > mca21_official > scraper_HIGH > screener > gemini
    """
    r = dict(db)

    def fill(field: str, value: Any):
        if _is_empty(r.get(field)) and not _is_empty(value):
            r[field] = value

    # ── RBI (highest authority: registration, category, location) ──
    if rbi:
        fill('rbi_registration_number', rbi.get('rbi_registration_number'))
        fill('rbi_category',            rbi.get('rbi_category'))
        fill('hq_state',                rbi.get('hq_state'))
        fill('hq_location',             rbi.get('hq_location'))

    # ── MCA21 (authoritative for registration, incorporation, capital) ──
    if mca21:
        fill('cin',                       mca21.get('cin'))
        fill('established_year',          mca21.get('established_year'))
        fill('hq_state',                  mca21.get('hq_state'))
        fill('authorized_capital_lakhs',  mca21.get('authorized_capital_lakhs'))
        fill('paid_up_capital_lakhs',     mca21.get('paid_up_capital_lakhs'))
        fill('is_listed',                 mca21.get('is_listed'))
        # company_status always written when MCA21 provided it (may have changed)
        if mca21.get('company_status'):
            r['company_status'] = mca21['company_status']
        # propagate tracking fields regardless of null rules
        if mca21.get('mca21_status'):
            r['mca21_status'] = mca21['mca21_status']

    # ── Screener (financial data — listed companies) ────────────
    if screener:
        fill('aum_crores',        screener.get('aum_crores'))
        fill('last_year_revenue', screener.get('last_year_revenue'))
        fill('is_listed',         screener.get('is_listed'))
        fill('stock_symbol',      screener.get('stock_symbol'))
        fill('established_year',  screener.get('established_year'))

    # ── Scraper (contact + products + states — direct from website) ──
    if scraper:
        conf = scraper.get('_confidence', {})
        def c(f): return conf.get(f, {}).get('confidence', 'LOW')

        for scraper_field, db_field, min_confs in [
            ('phone',            'phone',            ('HIGH', 'MEDIUM')),
            ('email',            'email',            ('HIGH', 'MEDIUM')),
            ('hq_state',         'hq_state',         ('HIGH', 'MEDIUM')),
            ('hq_city',          'hq_location',      ('HIGH', 'MEDIUM')),
            ('established_year', 'established_year', ('HIGH', 'MEDIUM')),
            ('employee_count',   'employee_count',   ('HIGH', 'MEDIUM')),
            ('branch_count',     'branch_count',     ('HIGH', 'MEDIUM')),
        ]:
            val = scraper.get(scraper_field)
            if not _is_empty(val) and c(scraper_field) in min_confs:
                fill(db_field, val)

        # Loan tags → primary_loan_segments
        tags = scraper.get('loan_tags', [])
        if tags:
            valid = [t for t in tags if t in VALID_PRODUCTS]
            fill('primary_loan_segments', valid or None)

        # Operating states
        raw_states = scraper.get('operating_states', [])
        if raw_states and not isinstance(raw_states, str):
            valid_states = [
                _STATE_LOWER.get(s.lower(), s) for s in raw_states
                if s == 'PAN_INDIA' or s in _STATE_SET or s.lower() in _STATE_LOWER
            ]
            if 'PAN_INDIA' in valid_states:
                fill('operating_states', sorted(ALL_INDIA_STATES))
                fill('pan_india', True)
                fill('operating_intensity', 'Pan India')
            else:
                fill('operating_states', [s for s in valid_states if s != 'PAN_INDIA'])
                if len(valid_states) >= 20:
                    fill('pan_india', True)
                    fill('operating_intensity', 'Pan India')

        # RBI registration from scraper — reject CIN numbers (MCA format)
        # CIN: L/U + 5digits + 2-letter state + 4-digit year + 3-letter type + 6digits
        # RBI CoR: "05.01915", "N.13.02437", "DNBS.XXX"
        reg = scraper.get('rbi_registration')
        if reg and any(c2.isdigit() for c2 in str(reg)):
            _cin_pattern = re.compile(r'^[LUu]\d{5}[A-Z]{2}\d{4}[A-Z]{3}\d{6}$')
            if not _cin_pattern.match(str(reg).strip()):
                fill('rbi_registration_number', reg)

        # Low-confidence fallbacks for contact fields
        for sf, df in [('phone', 'phone'), ('email', 'email'), ('hq_state', 'hq_state')]:
            fill(df, scraper.get(sf))

    # ── Recompute aum_category ─────────────────────────────────
    aum = r.get('aum_crores')
    if aum is not None:
        try:
            f = float(aum)
            r['aum_category'] = ('Micro' if f < 500 else
                                 'Small' if f < 5000 else
                                 'Mid'   if f < 50000 else 'Large')
        except (TypeError, ValueError):
            pass

    # ── Recompute quality_score ───────────────────────────────
    r['quality_score']      = _compute_quality(r)
    r['data_source']        = 'scraper+gemini'  # keep existing enum value
    r['extraction_status']  = 'success'

    return r


# ─────────────────────────────────────────────────────────────
# MAIN LOOP
# ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Enrich lenders — no Gemini')
    parser.add_argument('--approve',      action='store_true',
                        help='Auto-approve lenders with quality_score >= min-score after enrichment')
    parser.add_argument('--approve-only', action='store_true',
                        help='Skip enrichment, just approve by existing score')
    parser.add_argument('--only',         type=str, default=None,
                        help='Process only lenders whose name contains this string')
    parser.add_argument('--type',         type=str, default=None,
                        help='Filter by company_type e.g. NBFC, "PSU Bank"')
    parser.add_argument('--missing-rbi',  action='store_true',
                        help='Only process lenders missing rbi_registration_number')
    parser.add_argument('--missing-aum',  action='store_true',
                        help='Only process lenders missing aum_crores')
    parser.add_argument('--skip-rbi',     action='store_true')
    parser.add_argument('--skip-mca21',   action='store_true',
                        help='Skip MCA21 enrichment (Zauba Corp)')
    parser.add_argument('--skip-screener',action='store_true')
    parser.add_argument('--skip-scraper', action='store_true')
    parser.add_argument('--min-score',    type=float, default=0.45,
                        help='Quality score threshold for approval (default 0.45)')
    parser.add_argument('--dry-run',      action='store_true',
                        help='No DB writes')
    parser.add_argument('--limit',        type=int, default=None,
                        help='Process only first N lenders (useful for testing)')
    parser.add_argument('--delay',        type=float, default=2.0,
                        help='Seconds between Screener requests (default 2.0)')
    args = parser.parse_args()

    if not DATABASE_URL:
        log.error('DATABASE_URL not set in .env')
        sys.exit(1)

    conn = get_db()

    # ── Approve-only mode ──────────────────────────────────────
    if args.approve_only:
        log.info('Approve-only mode: min_score=%.2f', args.min_score)
        _, approved = bulk_update_quality_and_approve(conn, args.min_score, args.dry_run)
        conn.close()
        print(f'\nApproved: {approved} lenders')
        return

    # ── Fetch lenders ──────────────────────────────────────────
    filters = {
        'only':        args.only,
        'type':        args.type,
        'missing_rbi': args.missing_rbi,
        'missing_aum': args.missing_aum,
        'limit':       args.limit,
    }
    lenders = fetch_lenders(conn, filters)
    log.info('Processing %d lenders', len(lenders))

    # ── Prefetch RBI list once ─────────────────────────────────
    if not args.skip_rbi:
        log.info('Loading RBI NBFC list...')
        rbi_list = load_rbi_nbfc_list()
        log.info('RBI list: %d entries', len(rbi_list))

    # ── HTTP sessions ──────────────────────────────────────────
    screener_session = requests.Session()

    if not args.skip_screener:
        try:
            screener_session.get('https://www.screener.in/', headers=_HEADERS, timeout=10)
        except Exception:
            pass

    # ── Enrichment loop ────────────────────────────────────────
    stats = {'updated': 0, 'no_change': 0, 'errors': 0, 'dry_run': 0,
             'mca21_enriched': 0, 'mca21_not_found': 0}
    is_nbfc_types = {'NBFC', 'NBFC-MFI', 'NBFC-D', 'NBFC-ND'}

    for i, lender in enumerate(lenders, 1):
        name   = lender['company_name']
        ctype  = lender.get('company_type', 'NBFC')
        website= lender.get('website', '')

        log.info('\n[%d/%d] %s (%s)', i, len(lenders), name, ctype)

        try:
            # Source 1: RBI (NBFCs only — banks aren't in the NBFC list)
            rbi_data = None
            if not args.skip_rbi and ctype in is_nbfc_types:
                rbi_data = rbi_lookup(name)
                if rbi_data:
                    log.info('  [rbi] found: %s', {k: v for k, v in rbi_data.items() if v})

            # Source 2: MCA21 via Zauba Corp (all company types)
            mca21_data = None
            if not args.skip_mca21:
                # Skip if CIN already in DB and company_status already known
                already_enriched = (
                    not _is_empty(lender.get('cin'))
                    and lender.get('mca21_status') == 'enriched'
                )
                if not already_enriched:
                    mca21_data = mca21_enrich(name, website=website)
                else:
                    log.debug('  [mca21] already enriched, skipping')

                # Track MCA21 outcome in run stats
                if mca21_data:
                    s = mca21_data.get('mca21_status', '')
                    if s == 'enriched':
                        stats['mca21_enriched'] += 1
                    elif s == 'not_found':
                        stats['mca21_not_found'] += 1

            # Source 3: Screener.in (financial data for listed companies)
            screener_data = None
            if not args.skip_screener:
                time.sleep(args.delay)
                screener_data = screener_enrich(name, screener_session)

            # Source 4: Website scraper
            scraper_data = None
            if not args.skip_scraper and website:
                time.sleep(1.0)
                scraper_data = scraper_enrich(name, website)

            # Merge
            enriched = merge(lender, rbi_data, mca21_data, screener_data, scraper_data)

            # Log delta (what changed)
            changed = []
            for field in ['rbi_registration_number', 'rbi_category', 'aum_crores',
                          'last_year_revenue', 'hq_state', 'hq_location',
                          'primary_loan_segments', 'operating_states', 'phone', 'email',
                          'established_year', 'is_listed',
                          'cin', 'company_status', 'authorized_capital_lakhs',
                          'paid_up_capital_lakhs']:
                old = lender.get(field)
                new = enriched.get(field)
                if _is_empty(old) and not _is_empty(new):
                    changed.append(f'{field}={repr(new)[:40]}')
            if changed:
                log.info('  [delta] filled: %s', ' | '.join(changed))
            log.info('  [score] %.2f -> %.2f',
                     lender.get('quality_score') or 0, enriched['quality_score'])

            # Write to DB
            action = upsert_enriched(conn, enriched, dry_run=args.dry_run)
            stats[action] = stats.get(action, 0) + 1

        except KeyboardInterrupt:
            log.warning('\nInterrupted at %d/%d', i, len(lenders))
            break
        except Exception as exc:
            log.error('  [error] %s: %s', name, exc)
            stats['errors'] += 1

    # ── Auto-approve ───────────────────────────────────────────
    newly_approved = 0
    if args.approve and not args.dry_run:
        log.info('\nRunning auto-approve (min_score=%.2f)...', args.min_score)
        _, newly_approved = bulk_update_quality_and_approve(conn, args.min_score, args.dry_run)

    conn.close()

    # ── Summary ────────────────────────────────────────────────
    print(f'\n{"="*60}')
    print(f'ENRICHMENT COMPLETE')
    print(f'  Lenders processed  : {len(lenders)}')
    print(f'  Updated            : {stats.get("updated", 0)}')
    print(f'  No change          : {stats.get("no_change", 0)}')
    print(f'  Errors             : {stats.get("errors", 0)}')
    print(f'  MCA21 enriched     : {stats.get("mca21_enriched", 0)}')
    print(f'  MCA21 not found    : {stats.get("mca21_not_found", 0)}')
    if args.dry_run:
        print('  (DRY RUN -- no DB writes)')
    if args.approve:
        print(f'  Newly approved     : {newly_approved}')
    print(f'{"="*60}')


if __name__ == '__main__':
    main()
